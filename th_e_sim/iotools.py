# -*- coding: utf-8 -*-
"""
    th-e-sim.iotools
    ~~~~~


"""
import os
import copy
import logging
import pandas as pd
import matplotlib.pyplot as plt

logger = logging.getLogger(__name__)


# noinspection PyPackageRequirements
def write_excel(settings, summary, validations):
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Border, Font, Side

        border_side = Side(border_style=None)
        border = Border(top=border_side,
                        right=border_side,
                        bottom=border_side,
                        left=border_side)

        summary_file = os.path.join(settings.get('General', 'data_dir', fallback='data'), 'summary.xlsx')
        summary_book = Workbook()
        summary_writer = pd.ExcelWriter(summary_file, engine='openpyxl', engine_kwargs={'encoding': 'utf-8-sig'})
        summary_writer.book = summary_book
        summary.to_excel(summary_writer, sheet_name='Summary', float_format="%.2f", encoding='utf-8-sig')
        summary_book.remove_sheet(summary_book.active)
        summary_book.active = 0

        for validation_key, validation in validations.items():
            validation.to_excel(summary_writer, sheet_name=validation_key, encoding='utf-8-sig')

        # Set column width and header coloring
        for summary_sheet in summary_book:
            summary_sheet.delete_rows(3, 1)
            summary_index_width = 0
            for summary_row in summary_sheet:
                summary_row[0].border = border
                summary_index_width = max(summary_index_width, len(str(summary_row[0].value)))

            summary_sheet.column_dimensions[get_column_letter(1)].width = summary_index_width + 2

            summary_header = len(summary.columns.levels)
            summary_header_font = Font(name="Calibri Light", size=12, color='333333')
            for summary_column in range(1, len(summary_sheet[summary_header])):
                summary_column_width = len(str(summary_sheet[summary_header][summary_column].value))

                summary_sheet[summary_header][summary_column].border = border
                for summary_header_row in range(1, summary_header):
                    summary_sheet[summary_header_row][summary_column].font = summary_header_font
                    summary_sheet[summary_header_row][summary_column].border = border
                summary_sheet[summary_header][summary_column].border = border
                summary_sheet.column_dimensions[get_column_letter(summary_column+1)].width = summary_column_width + 2

        summary_book.save(summary_file)
        summary_writer.close()

    except ImportError:
        pass


def write_csv(system, data, file):
    system_dir = system.configs['General']['data_dir']
    database = copy.deepcopy(system._database)
    database.dir = system_dir
    # database.format = '%Y%m%d'
    database.enabled = True
    data_file = os.path.join(database.dir, file + '.csv')
    data_dir = os.path.dirname(data_file)

    if not os.path.isdir(data_dir):
        os.makedirs(data_dir, exist_ok=True)

    data.to_csv(data_file,
                sep=database.separator,
                decimal=database.decimal,
                encoding='utf-8-sig')


def print_distributions(features, path=''):
    # Desired number of bins in each plot
    bin_num = 100
    for feature in features.columns:  # create 100 equal space bin vals per feat.
        bins = []
        bin_domain = features[feature].max() - features[feature].min()
        bin_step = bin_domain / bin_num

        counter = features[feature].min()
        for i in range(bin_num):
            bins.append(counter)
            counter = counter + bin_step

        # Add the last value of the counter
        bins.append(counter)

        bin_values, bins, patches = plt.hist(features[feature], bins=bins)
        count_range = max(bin_values) - min(bin_values)
        sorted_values = list(bin_values)
        sorted_values.sort(reverse=True)

        # Scale plots by step through sorted bins
        for i in range(len(sorted_values) - 1):
            if abs(sorted_values[i] - sorted_values[i + 1]) / count_range < 0.80:
                continue
            else:
                plt.ylim([0, sorted_values[i + 1] + 10])
                break

        # Save histogram to appropriate folder
        path_dist = os.path.join(path, 'dist')
        path_file = os.path.join(path_dist, '{}.png'.format(feature))
        if not os.path.isdir(path_dist):
            os.makedirs(path_dist, exist_ok=True)

        plt.title(r'Histogram of '+feature)
        plt.savefig(path_file)
        plt.clf()


# noinspection PyPackageRequirements
def print_boxplot(system, data, index, column, file, label='', title='', colors=None, **kwargs):
    import seaborn as sns

    plt.figure()
    plot_fliers = dict(marker='o', markersize=3, markerfacecolor='none', markeredgecolor='lightgrey')
    plot_colors = colors if colors is not None else index.nunique()
    plot_palette = sns.light_palette('#0069B4', n_colors=plot_colors, reverse=True)
    plot = sns.boxplot(x=index, y=column,
                       data=data,
                       palette=plot_palette,
                       flierprops=plot_fliers,
                       # showfliers=False,
                       **kwargs)
    plot.set(xlabel=label, ylabel='Error [W]', title=title)
    plt.show(block=False)

    plot_file = os.path.join(system.configs['General']['data_dir'], file + '.png')
    plot_dir = os.path.dirname(plot_file)

    if not os.path.isdir(plot_dir):
        os.makedirs(plot_dir, exist_ok=True)

    plot.figure.savefig(plot_file)
