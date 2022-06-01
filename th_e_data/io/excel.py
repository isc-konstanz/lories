# -*- coding: utf-8 -*-
"""
    th-e-sim.io.excel
    ~~~~~~~~~~~~~~~~~~~~~~


"""
import os
import logging
import warnings
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, Side
from tables import NaturalNameWarning
from copy import copy

warnings.filterwarnings('ignore', category=NaturalNameWarning)
logger = logging.getLogger(__name__)


def write_excel(settings, summary, validations):
    border_side = Side(border_style=None)
    border = Border(top=border_side,
                    right=border_side,
                    bottom=border_side,
                    left=border_side)

    summary_file = os.path.join(settings.get('General', 'data_dir', fallback='data'), 'summary.xlsx')
    summary_book = Workbook()
    summary_writer = pd.ExcelWriter(summary_file, engine='openpyxl')
    summary_writer.book = summary_book
    summary.to_excel(summary_writer, sheet_name='Summary', float_format="%.2f", encoding='utf-8-sig')
    summary_book.remove_sheet(summary_book.active)
    summary_book.active = 0

    for validation_key, validation in validations.items():
        validation.to_excel(summary_writer, sheet_name=validation_key, encoding='utf-8-sig')
        validation_sheet = summary_book[validation_key]
        for validation_column in range(1, len(validation_sheet[1])):
            validation_column_value = validation_sheet[1][validation_column].value
            if validation_column_value is not None:
                validation_column_width = len(validation_column_value) + 2
                validation_sheet.column_dimensions[get_column_letter(validation_column + 1)].width = \
                    validation_column_width
            validation_sheet[1][validation_column].border = border

    # Set column width and header coloring
    for summary_sheet in summary_book:
        if summary_sheet.title == 'Summary':
            summary_sheet.delete_rows(3, 1)

        summary_index_width = 0
        for summary_row in summary_sheet:
            summary_row[0].border = border
            summary_index_width = max(summary_index_width, len(str(summary_row[0].value)))

        summary_sheet.column_dimensions[get_column_letter(1)].width = summary_index_width + 2

        summary_header = len(summary.columns.levels)
        summary_header_font = Font(name="Calibri Light", size=12, color='333333')
        for summary_column in range(len(summary_sheet[summary_header])):
            for summary_header_row in range(1, summary_header):
                if '\n' in str(summary_sheet[summary_header_row][summary_column].value):
                    summary_header_alignment = copy(summary_sheet[summary_header_row][summary_column].alignment)
                    summary_header_alignment.wrapText = True
                    # summary_header_alignment.vertical = 'center'
                    # summary_header_alignment.horizontal = 'center'
                    summary_sheet[summary_header_row][summary_column].alignment = summary_header_alignment
                    summary_sheet.row_dimensions[summary_header_row].height = 33

                summary_sheet[summary_header_row][summary_column].font = summary_header_font
                summary_sheet[summary_header_row][summary_column].border = border

            summary_sheet[summary_header][summary_column].border = border
            if summary_header == 1:
                summary_header_width = len(str(summary_sheet[summary_header][summary_column].value))
                summary_sheet.column_dimensions[get_column_letter(summary_column+1)].width = summary_header_width + 2

    summary_book.save(summary_file)
    summary_writer.close()
